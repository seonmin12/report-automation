"""
Excel 검증 리포트 생성 모듈.

openpyxl을 사용해 실무형 엑셀 리포트를 만든다. 시트 구성:
1. 요약 (전체 이슈 건수, 실적 차이 개요)
2. 실적비교 (compare.py 결과 - 실시간 실적 공지 대응)
3. 요금제미매핑이슈 (validate.py 결과 4종 통합 - 요금제 미매핑 현황 대응)
4. 오류상세 (실적비교 + 매핑이슈를 하나로 합친 전체 오류 목록)
5. 사업자별누적가입자 (aggregate.py 결과 - 과기부 제출자료 대응)

이상 항목은 조건부 서식(빨간 배경 등)으로 강조한다.
"""

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from config import (
    COL_OPERATOR_CODE,
    COL_OPERATOR_NAME,
    COL_PRODUCT_CODE,
    COL_PRODUCT_NAME,
    COL_VALIDATION_STATUS,
    DUMMY_OPERATORS,
    STATUS_NEEDS_REVIEW,
)

from .aggregate import COL_CUMULATIVE_COUNT
from .compare import COL_DIFF, COL_DIFF_PCT, COL_ISSUE_FLAG, COL_SOURCE, SOURCE_BOTH
from .validate import COL_ISSUE_DETAIL, COL_ISSUE_TYPE, build_issue_detail_table

ISSUE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

DEFAULT_COLUMN_WIDTH = 16

ISSUE_TYPE_COMPARE_MISMATCH = "포털-RAW 실적 불일치"

_OPERATOR_NAME_MAP = {op[COL_OPERATOR_CODE]: op[COL_OPERATOR_NAME] for op in DUMMY_OPERATORS}


def write_dataframe_to_sheet(
    ws,
    df: pd.DataFrame,
    highlight_col: str | None = None,
    highlight_value=None,
    highlight_all: bool = False,
):
    """DataFrame을 워크시트에 쓰고 헤더 스타일 적용.

    highlight_all=True면 모든 데이터 행을 강조한다 (표 전체가 이슈 목록인 경우).
    highlight_col이 주어지면 해당 컬럼 값을 기준으로 강조한다.
    highlight_value를 지정하면 그 값과 일치하는 행만, 지정하지 않으면 참(True)인 행을 강조한다.
    """
    columns = list(df.columns)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    highlight_col_idx = None
    if highlight_col is not None and highlight_col in columns:
        highlight_col_idx = columns.index(highlight_col) + 1

    if highlight_all or highlight_col_idx is not None:
        for row_idx in range(2, ws.max_row + 1):
            should_highlight = highlight_all
            if highlight_col_idx is not None:
                cell_value = ws.cell(row=row_idx, column=highlight_col_idx).value
                if highlight_value is not None:
                    should_highlight = should_highlight or (cell_value == highlight_value)
                else:
                    should_highlight = should_highlight or bool(cell_value)
            if should_highlight:
                for cell in ws[row_idx]:
                    cell.fill = ISSUE_FILL

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            DEFAULT_COLUMN_WIDTH
        )


def build_error_detail_df(compare_df: pd.DataFrame, validation_results: dict) -> pd.DataFrame:
    """실적비교 이상 항목 + 매핑 이슈 4종을 하나의 오류 목록으로 통합."""
    rows = []

    if not compare_df.empty:
        portal_name_col = f"{COL_PRODUCT_NAME}_포털"
        raw_name_col = f"{COL_PRODUCT_NAME}_RAW"
        for _, row in compare_df.loc[compare_df[COL_ISSUE_FLAG]].iterrows():
            if row[COL_SOURCE] == SOURCE_BOTH:
                detail = f"순증건수 차이 {row[COL_DIFF]}건 (차이율 {row[COL_DIFF_PCT]}%)"
            else:
                detail = row[COL_SOURCE]

            product_name = row[portal_name_col]
            if pd.isna(product_name):
                product_name = row[raw_name_col]

            rows.append(
                {
                    COL_ISSUE_TYPE: ISSUE_TYPE_COMPARE_MISMATCH,
                    COL_OPERATOR_CODE: row[COL_OPERATOR_CODE],
                    COL_OPERATOR_NAME: row[COL_OPERATOR_NAME],
                    COL_PRODUCT_CODE: row[COL_PRODUCT_CODE],
                    COL_PRODUCT_NAME: product_name,
                    COL_ISSUE_DETAIL: detail,
                }
            )

    for _, row in build_issue_detail_table(validation_results).iterrows():
        rows.append(
            {
                COL_ISSUE_TYPE: row[COL_ISSUE_TYPE],
                COL_OPERATOR_CODE: row[COL_OPERATOR_CODE],
                COL_OPERATOR_NAME: _OPERATOR_NAME_MAP.get(row[COL_OPERATOR_CODE], "-"),
                COL_PRODUCT_CODE: row[COL_PRODUCT_CODE],
                COL_PRODUCT_NAME: row[COL_PRODUCT_NAME],
                COL_ISSUE_DETAIL: row[COL_ISSUE_DETAIL],
            }
        )

    columns = [
        COL_ISSUE_TYPE,
        COL_OPERATOR_CODE,
        COL_OPERATOR_NAME,
        COL_PRODUCT_CODE,
        COL_PRODUCT_NAME,
        COL_ISSUE_DETAIL,
    ]
    return pd.DataFrame(rows, columns=columns)


def _write_summary_sheet(
    ws,
    as_of_date: date,
    compare_df: pd.DataFrame,
    error_detail_df: pd.DataFrame,
    operator_summary_df: pd.DataFrame,
):
    ws.append(["실적 검증 요약"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"검증 기준일: {as_of_date.strftime('%Y-%m-%d')}"])
    ws.append([])

    ws.append(["항목", "건수"])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    compare_total = len(compare_df)
    compare_issue_count = int(compare_df[COL_ISSUE_FLAG].sum()) if not compare_df.empty else 0
    ws.append(["실적 비교 대상 (사업자/상품 조합)", compare_total])
    ws.append(["  - 정상", compare_total - compare_issue_count])
    ws.append(["  - 확인필요", compare_issue_count])
    ws.append([])

    ws.append(["전체 오류 건수 (오류상세 시트 기준)", len(error_detail_df)])
    if not error_detail_df.empty:
        for issue_type, count in error_detail_df[COL_ISSUE_TYPE].value_counts().items():
            ws.append([f"  - {issue_type}", int(count)])
    ws.append([])

    total_subscribers = (
        int(operator_summary_df[COL_CUMULATIVE_COUNT].sum()) if not operator_summary_df.empty else 0
    )
    ws.append(["사업자 전체 누적가입자수", total_subscribers])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16


def build_excel_report(
    as_of_date: date,
    compare_df: pd.DataFrame,
    validation_results: dict,
    operator_summary_df: pd.DataFrame,
    output_path,
):
    """5개 시트로 구성된 검증 리포트 엑셀 생성."""
    error_detail_df = build_error_detail_df(compare_df, validation_results)

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "요약"
    _write_summary_sheet(summary_ws, as_of_date, compare_df, error_detail_df, operator_summary_df)

    compare_ws = wb.create_sheet("실적비교")
    write_dataframe_to_sheet(compare_ws, compare_df, highlight_col=COL_ISSUE_FLAG)

    mapping_issue_ws = wb.create_sheet("요금제미매핑이슈")
    write_dataframe_to_sheet(
        mapping_issue_ws, build_issue_detail_table(validation_results), highlight_all=True
    )

    error_detail_ws = wb.create_sheet("오류상세")
    write_dataframe_to_sheet(error_detail_ws, error_detail_df, highlight_all=True)

    operator_ws = wb.create_sheet("사업자별누적가입자")
    write_dataframe_to_sheet(
        operator_ws,
        operator_summary_df,
        highlight_col=COL_VALIDATION_STATUS,
        highlight_value=STATUS_NEEDS_REVIEW,
    )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
